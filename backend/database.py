"""
Nexus Puppy Flow — Database Layer
SQLite database with all tables for the puppy tracking system.
"""

import sqlite3
import os
import json
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "nexus_puppy.db")


def get_connection():
    """Get a database connection with row_factory set."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """Create all tables if they don't exist."""
    conn = get_connection()
    cursor = conn.cursor()

    # Puppies / profiles
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS puppies (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            gender TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'puppy',
            color TEXT DEFAULT '',
            avatar TEXT DEFAULT '🐶',
            avatar_bg TEXT DEFAULT 'rgba(77,171,247,0.15)',
            avatar_color TEXT DEFAULT '#4dabf7',
            notes TEXT DEFAULT '',
            birth_date TEXT DEFAULT NULL
        )
    """)

    # Weight records
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS weights (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            puppy_id TEXT NOT NULL,
            date TEXT NOT NULL,
            value REAL NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (puppy_id) REFERENCES puppies(id)
        )
    """)
    cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_weights_puppy_date
        ON weights(puppy_id, date)
    """)

    # Feeding records
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS feedings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            time_key TEXT NOT NULL,
            block_a INTEGER DEFAULT 0,
            block_b INTEGER DEFAULT 0,
            timestamp INTEGER,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_feedings_date
        ON feedings(date, time_key)
    """)

    # Medical events
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS medical_events (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            date TEXT NOT NULL,
            type TEXT DEFAULT 'checkup',
            status TEXT DEFAULT 'pending',
            for_puppies INTEGER DEFAULT 1
        )
    """)

    # Custom events (user-created)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS custom_events (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            date TEXT NOT NULL,
            type TEXT DEFAULT 'checkup',
            for_puppies INTEGER DEFAULT 0
        )
    """)

    # Medical event status (done/pending per user)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS medical_status (
            event_id TEXT PRIMARY KEY,
            status TEXT DEFAULT 'pending'
        )
    """)

    # Puppy notes
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS puppy_notes (
            puppy_id TEXT PRIMARY KEY,
            notes TEXT DEFAULT ''
        )
    """)

    # Blanquita meals tracking
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS blanquita_meals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            time_str TEXT NOT NULL,
            served INTEGER DEFAULT 0,
            portion REAL DEFAULT 0,
            notes TEXT DEFAULT '',
            timestamp INTEGER
        )
    """)
    cursor.execute("""
        CREATE INDEX IF NOT EXISTS idx_blanquita_meals_date
        ON blanquita_meals(date, time_str)
    """)

    # Blanquita reminders config
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS blanquita_reminders (
            id INTEGER PRIMARY KEY DEFAULT 1,
            enabled INTEGER DEFAULT 0,
            minutes_before INTEGER DEFAULT 5
        )
    """)
    # Ensure default row exists
    cursor.execute("""
        INSERT OR IGNORE INTO blanquita_reminders (id, enabled, minutes_before)
        VALUES (1, 0, 5)
    """)

    # Migration tracking
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS migration_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            migrated_at TEXT DEFAULT (datetime('now')),
            data_type TEXT,
            records_count INTEGER
        )
    """)

    conn.commit()
    conn.close()
    print("✅ Database initialized successfully")


# ─── CRUD: Puppies ───

def get_puppies():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM puppies ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def seed_puppies():
    """Insert default puppy data if table is empty."""
    conn = get_connection()
    existing = conn.execute("SELECT COUNT(*) FROM puppies").fetchone()[0]
    if existing > 0:
        conn.close()
        return

    default_puppies = [
        {"id": "blanquita", "name": "Blanquita", "gender": "F", "role": "mother",
         "color": "Blanco", "avatar": "🐕", "avatar_bg": "rgba(46,204,113,0.15)",
         "avatar_color": "#2ecc71", "notes": "Madre de la camada. Semi-callejera del condominio."},
        {"id": "max", "name": "Max", "gender": "M", "role": "puppy",
         "color": "Patrón Steel", "avatar": "🐶", "avatar_bg": "rgba(77,171,247,0.15)",
         "avatar_color": "#4dabf7", "notes": "Líder del Bloque A. Tranquilo.", "birth_date": "2026-05-23"},
        {"id": "steel", "name": "Steel", "gender": "M", "role": "puppy",
         "color": "Patrón Steel", "avatar": "🐶", "avatar_bg": "rgba(77,171,247,0.15)",
         "avatar_color": "#4dabf7", "notes": "Parte del Bloque A. Tranquilo como Max.", "birth_date": "2026-05-23"},
        {"id": "sydney", "name": "Sydney", "gender": "M", "role": "puppy",
         "color": "Marrón claro", "avatar": "🐶", "avatar_bg": "rgba(77,171,247,0.15)",
         "avatar_color": "#4dabf7", "notes": "De quien se enamora Max. Bloque A.", "birth_date": "2026-05-23"},
        {"id": "arturo", "name": "Arturo", "gender": "M", "role": "puppy",
         "color": "Marrón oscuro", "avatar": "🐶", "avatar_bg": "rgba(77,171,247,0.15)",
         "avatar_color": "#4dabf7", "notes": "Bloque B. Es de los más grandes y fuertes.", "birth_date": "2026-05-23"},
        {"id": "travieso", "name": "Travieso", "gender": "M", "role": "puppy",
         "color": "Pequeño dominante", "avatar": "🐕", "avatar_bg": "rgba(224,184,92,0.15)",
         "avatar_color": "#e0b85c", "birth_date": "2026-05-23",
         "notes": "EL MÁS PEQUEÑO. Prioridad máxima en alimentación. Bloque B — siempre en tetas traseras de Blanquita."},
        {"id": "chana", "name": "Chana", "gender": "F", "role": "puppy",
         "color": "Blanco con manchas", "avatar": "🐩", "avatar_bg": "rgba(232,125,158,0.15)",
         "avatar_color": "#e87d9e", "notes": "Bloque B. Fuerte, come bien.", "birth_date": "2026-05-23"},
        {"id": "alofoka", "name": "Alofoka", "gender": "F", "role": "puppy",
         "color": "Gris claro", "avatar": "🐩", "avatar_bg": "rgba(232,125,158,0.15)",
         "avatar_color": "#e87d9e", "notes": "Bloque A. Tranquila.", "birth_date": "2026-05-23"},
        {"id": "rodotesa", "name": "Rodotesa", "gender": "F", "role": "puppy",
         "color": "Marrón claro", "avatar": "🐩", "avatar_bg": "rgba(232,125,158,0.15)",
         "avatar_color": "#e87d9e", "notes": "Le gusta rodar. Bloque B.", "birth_date": "2026-05-23"},
    ]

    for p in default_puppies:
        conn.execute("""INSERT OR IGNORE INTO puppies
            (id, name, gender, role, color, avatar, avatar_bg, avatar_color, notes, birth_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                     (p["id"], p["name"], p["gender"], p["role"], p["color"],
                      p["avatar"], p["avatar_bg"], p["avatar_color"],
                      p.get("notes", ""), p.get("birth_date")))
    conn.commit()
    conn.close()
    print(f"✅ Seeded {len(default_puppies)} puppies")


# ─── CRUD: Weights ───

def get_weights(puppy_id=None):
    conn = get_connection()
    if puppy_id:
        rows = conn.execute(
            "SELECT * FROM weights WHERE puppy_id = ? ORDER BY date", (puppy_id,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM weights ORDER BY date").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_weight(puppy_id: str, date: str, value: float):
    conn = get_connection()
    conn.execute(
        "INSERT INTO weights (puppy_id, date, value) VALUES (?, ?, ?)",
        (puppy_id, date, value),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


def delete_weight(weight_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM weights WHERE id = ?", (weight_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ─── CRUD: Feedings ───

def get_feedings(date=None):
    conn = get_connection()
    if date:
        rows = conn.execute(
            "SELECT * FROM feedings WHERE date = ? ORDER BY time_key", (date,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM feedings ORDER BY date DESC, time_key").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def upsert_feeding(date: str, time_key: str, block_a: bool = False, block_b: bool = False):
    conn = get_connection()
    existing = conn.execute(
        "SELECT id FROM feedings WHERE date = ? AND time_key = ?", (date, time_key)
    ).fetchone()

    if existing:
        if block_a:
            conn.execute("UPDATE feedings SET block_a = 1, timestamp = ? WHERE id = ?",
                         (int(datetime.now().timestamp()), existing["id"]))
        if block_b:
            conn.execute("UPDATE feedings SET block_b = 1, timestamp = ? WHERE id = ?",
                         (int(datetime.now().timestamp()), existing["id"]))
    else:
        conn.execute(
            """INSERT INTO feedings (date, time_key, block_a, block_b, timestamp)
               VALUES (?, ?, ?, ?, ?)""",
            (date, time_key, 1 if block_a else 0, 1 if block_b else 0,
             int(datetime.now().timestamp())),
        )
    conn.commit()
    conn.close()
    return {"ok": True}


# ─── CRUD: Medical ───

def get_medical_events():
    conn = get_connection()
    base = conn.execute("SELECT * FROM medical_events ORDER BY date").fetchall()
    custom = conn.execute("SELECT * FROM custom_events ORDER BY date").fetchall()
    statuses = conn.execute("SELECT * FROM medical_status").fetchall()
    conn.close()

    status_map = {s["event_id"]: s["status"] for s in statuses}
    events = []
    for e in base:
        d = dict(e)
        d["status"] = status_map.get(d["id"], "pending")
        events.append(d)
    for e in custom:
        d = dict(e)
        d["status"] = status_map.get(d["id"], "pending")
        events.append(d)
    return events


def toggle_medical_event(event_id: str):
    conn = get_connection()
    existing = conn.execute(
        "SELECT status FROM medical_status WHERE event_id = ?", (event_id,)
    ).fetchone()
    if existing:
        new_status = "pending" if existing["status"] == "done" else "done"
        conn.execute("UPDATE medical_status SET status = ? WHERE event_id = ?",
                     (new_status, event_id))
    else:
        conn.execute("INSERT INTO medical_status (event_id, status) VALUES (?, 'done')",
                     (event_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


def add_custom_event(title: str, date: str, type_: str, description: str = ""):
    import hashlib
    event_id = "custom_" + hashlib.md5(f"{title}{date}{datetime.now().isoformat()}".encode()).hexdigest()[:12]
    conn = get_connection()
    conn.execute(
        "INSERT INTO custom_events (id, title, description, date, type) VALUES (?, ?, ?, ?, ?)",
        (event_id, title, description, date, type_),
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": event_id}


# ─── CRUD: Blanquita Meals ───

def get_blanquita_meals(date: str = None):
    conn = get_connection()
    if date:
        rows = conn.execute(
            "SELECT * FROM blanquita_meals WHERE date = ? ORDER BY time_str", (date,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM blanquita_meals ORDER BY date DESC, time_str"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def mark_blanquita_meal(date: str, time_str: str, portion: float = 0, notes: str = ""):
    conn = get_connection()
    existing = conn.execute(
        "SELECT id FROM blanquita_meals WHERE date = ? AND time_str = ?",
        (date, time_str),
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE blanquita_meals SET served = 1, portion = ?, notes = ?, timestamp = ? WHERE id = ?",
            (portion, notes, int(datetime.now().timestamp()), existing["id"]),
        )
    else:
        conn.execute(
            "INSERT INTO blanquita_meals (date, time_str, served, portion, notes, timestamp) VALUES (?, ?, 1, ?, ?, ?)",
            (date, time_str, portion, notes, int(datetime.now().timestamp())),
        )
    conn.commit()
    conn.close()
    return {"ok": True}


# ─── CRUD: Blanquita Reminders ───

def get_reminder_config():
    conn = get_connection()
    row = conn.execute("SELECT * FROM blanquita_reminders WHERE id = 1").fetchone()
    conn.close()
    if row:
        return {"enabled": bool(row["enabled"]), "minutes_before": row["minutes_before"]}
    return {"enabled": False, "minutes_before": 5}


def update_reminder_config(enabled: bool, minutes_before: int = 5):
    conn = get_connection()
    conn.execute(
        "UPDATE blanquita_reminders SET enabled = ?, minutes_before = ? WHERE id = 1",
        (1 if enabled else 0, minutes_before),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


# ─── Migration ───

def migrate_from_localstorage(data: dict) -> dict:
    """Import data from localStorage JSON format into SQLite."""
    conn = get_connection()
    counts = {}

    # Migrate weights
    if "weights" in data:
        w_count = 0
        for puppy_id, records in data["weights"].items():
            for rec in records:
                conn.execute(
                    "INSERT OR IGNORE INTO weights (puppy_id, date, value) VALUES (?, ?, ?)",
                    (puppy_id, rec["date"], rec["value"]),
                )
                w_count += 1
        counts["weights"] = w_count

    # Migrate feedings
    if "feedings" in data:
        f_count = 0
        for date_str, time_slots in data["feedings"].items():
            for time_key, slot in time_slots.items():
                if isinstance(slot, dict):
                    conn.execute(
                        """INSERT OR IGNORE INTO feedings (date, time_key, block_a, block_b, timestamp)
                           VALUES (?, ?, ?, ?, ?)""",
                        (date_str, time_key,
                         1 if slot.get("blockA") else 0,
                         1 if slot.get("blockB") else 0,
                         slot.get("timestamp", int(datetime.now().timestamp()))),
                    )
                    f_count += 1
        counts["feedings"] = f_count

    # Migrate medical status
    if "medicalStatus" in data:
        m_count = 0
        for event_id, status in data["medicalStatus"].items():
            conn.execute(
                "INSERT OR IGNORE INTO medical_status (event_id, status) VALUES (?, ?)",
                (event_id, status),
            )
            m_count += 1
        counts["medical_status"] = m_count

    # Migrate custom events
    if "customEvents" in data:
        c_count = 0
        for ev in data["customEvents"]:
            conn.execute(
                "INSERT OR IGNORE INTO custom_events (id, title, description, date, type) VALUES (?, ?, ?, ?, ?)",
                (ev["id"], ev["title"], ev.get("desc", ""), ev["date"], ev.get("type", "other")),
            )
            c_count += 1
        counts["custom_events"] = c_count

    # Migrate puppy notes
    if "puppyNotes" in data:
        n_count = 0
        for puppy_id, notes in data["puppyNotes"].items():
            conn.execute(
                "INSERT OR IGNORE INTO puppy_notes (puppy_id, notes) VALUES (?, ?)",
                (puppy_id, notes),
            )
            n_count += 1
        counts["puppy_notes"] = n_count

    # Migrate blanquita meals
    if "blanquitaMeals" in data:
        b_count = 0
        for date_str, meals in data["blanquitaMeals"].items():
            for time_str, meal in meals.items():
                if isinstance(meal, dict):
                    conn.execute(
                        """INSERT OR IGNORE INTO blanquita_meals (date, time_str, served, portion, notes, timestamp)
                           VALUES (?, ?, ?, ?, ?, ?)""",
                        (date_str, time_str,
                         1 if meal.get("served") else 0,
                         meal.get("portion", 0),
                         meal.get("notes", ""),
                         meal.get("timestamp", int(datetime.now().timestamp()))),
                    )
                    b_count += 1
        counts["blanquita_meals"] = b_count

    # Log migration
    for data_type, count in counts.items():
        conn.execute(
            "INSERT INTO migration_log (data_type, records_count) VALUES (?, ?)",
            (data_type, count),
        )

    conn.commit()
    conn.close()
    return {"ok": True, "counts": counts}


# ─── Stats ───

def get_stats():
    conn = get_connection()
    stats = {
        "total_puppies": conn.execute("SELECT COUNT(*) FROM puppies").fetchone()[0],
        "total_weights": conn.execute("SELECT COUNT(*) FROM weights").fetchone()[0],
        "total_feedings": conn.execute("SELECT COUNT(*) FROM feedings").fetchone()[0],
        "total_medical": conn.execute("SELECT COUNT(*) FROM medical_events").fetchone()[0],
        "total_custom_events": conn.execute("SELECT COUNT(*) FROM custom_events").fetchone()[0],
        "total_blanquita_meals": conn.execute("SELECT COUNT(*) FROM blanquita_meals").fetchone()[0],
        "migration_log": conn.execute("SELECT * FROM migration_log ORDER BY migrated_at DESC").fetchall(),
        "db_size_kb": round(os.path.getsize(DB_PATH) / 1024, 1) if os.path.exists(DB_PATH) else 0,
    }
    conn.close()
    stats["migration_log"] = [dict(r) for r in stats["migration_log"]]
    return stats


# ─── Export ───

def export_all_json() -> str:
    conn = get_connection()
    data = {
        "puppies": [dict(r) for r in conn.execute("SELECT * FROM puppies").fetchall()],
        "weights": [dict(r) for r in conn.execute("SELECT * FROM weights").fetchall()],
        "feedings": [dict(r) for r in conn.execute("SELECT * FROM feedings ORDER BY date, time_key").fetchall()],
        "medical_events": [dict(r) for r in conn.execute("SELECT * FROM medical_events ORDER BY date").fetchall()],
        "custom_events": [dict(r) for r in conn.execute("SELECT * FROM custom_events ORDER BY date").fetchall()],
        "medical_status": [dict(r) for r in conn.execute("SELECT * FROM medical_status").fetchall()],
        "puppy_notes": [dict(r) for r in conn.execute("SELECT * FROM puppy_notes").fetchall()],
        "blanquita_meals": [dict(r) for r in conn.execute("SELECT * FROM blanquita_meals ORDER BY date, time_str").fetchall()],
        "exported_at": datetime.now().isoformat(),
        "version": "2.0",
    }
    conn.close()
    return json.dumps(data, ensure_ascii=False, indent=2)


def export_excel(filepath: str):
    """Export all data to an Excel workbook."""
    from openpyxl import Workbook
    wb = Workbook()

    # Puppies sheet
    ws = wb.active
    ws.title = "Perfiles"
    ws.append(["ID", "Nombre", "Género", "Rol", "Color", "Avatar", "Notas", "Fecha Nac."])
    for p in get_puppies():
        ws.append([p["id"], p["name"], p["gender"], p["role"], p["color"],
                   p["avatar"], p["notes"], p.get("birth_date", "")])

    # Weights sheet
    ws2 = wb.create_sheet("Pesos")
    ws2.append(["ID", "Cachorro ID", "Fecha", "Valor (g)"])
    for w in get_weights():
        ws2.append([w["id"], w["puppy_id"], w["date"], w["value"]])

    # Feedings sheet
    ws3 = wb.create_sheet("Alimentaciones")
    ws3.append(["ID", "Fecha", "Horario", "Bloque A", "Bloque B"])
    for f in get_feedings():
        ws3.append([f["id"], f["date"], f["time_key"],
                    "✅" if f["block_a"] else "⬜",
                    "✅" if f["block_b"] else "⬜"])

    # Medical sheet
    ws4 = wb.create_sheet("Eventos Médicos")
    ws4.append(["ID", "Título", "Descripción", "Fecha", "Tipo", "Estado"])
    for e in get_medical_events():
        ws4.append([e["id"], e["title"], e.get("description", ""),
                    e["date"], e["type"], e["status"]])

    # Blanquita meals sheet
    ws5 = wb.create_sheet("Comidas Blanquita")
    ws5.append(["ID", "Fecha", "Horario", "Servido", "Porción (g)", "Notas"])
    for b in get_blanquita_meals():
        ws5.append([b["id"], b["date"], b["time_str"],
                    "✅" if b["served"] else "⬜",
                    b["portion"], b["notes"]])

    # Stats sheet
    ws6 = wb.create_sheet("Estadísticas")
    stats = get_stats()
    ws6.append(["Métrica", "Valor"])
    for k, v in stats.items():
        if not isinstance(v, list):
            ws6.append([k.replace("_", " ").title(), v])

    wb.save(filepath)
    return filepath
